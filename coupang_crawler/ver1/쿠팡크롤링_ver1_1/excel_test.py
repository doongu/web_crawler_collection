from openpyxl import load_workbook, Workbook
import test
# data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook("C:/Users/Hellllo/Desktop/크몽/쿠팡 크롤링/쿠팡크롤링_ver1/test.xlsx", data_only=True)
# 시트 이름으로 불러오기
load_ws = load_wb['Sheet1']
index_answer = 0
check_name_list = []
def check_name_alg(input_row_c):
    for i in range(len(input_row_c)):
        if input_row_c[i] == "*":
            check_name_list.append([input_row_c[:i], input_row_c[i+1:]])
            break
    else:
        check_name_list.append([input_row_c])

excel_list = []
temp_excel_list = []
    # 셀 주소로 값 출력
for i in range(55, 963):
    input_row_b = str(load_ws.cell(i, 2).value)#상품명
    input_row_c = str(load_ws.cell(i, 3).value) #수량
    check_name_alg(input_row_c)
    input_row_d = str(load_ws.cell(i, 4).value) #가격
    # print(input_row_d)
    # print(type(input_row_d))
    if (input_row_b == "None" and input_row_c ==  "None") or (input_row_d == "None"):
        if len(temp_excel_list) != 0:
            excel_list.append(temp_excel_list)
            temp_excel_list = []
        continue
    else:
        temp_excel_list.append([input_row_b, input_row_b +  " " +input_row_c , int(input_row_d.replace(",", "")) ])

suma= 0
for i in range(len(excel_list)):
    print(excel_list[i])
    suma += len(excel_list[i])
print(suma)
excel_index = 0
breaker = 0

for i in range(len( check_name_list)):
    print(check_name_list[i])


while True:
    if breaker == 25:
        break
    # #가격을 리스트형태로 받는다.
    input_list = (test.select_data(excel_list[excel_index], check_name_list))

    #엑셀 열고
    write_wb =  Workbook()
    write_ws = write_wb.active

    # input_list[i][0]  i번째 제품 가격
    # input_list[i][1]  i번째 제품 url
    print(excel_list)
    for i in range(0, len(input_list)):
        write_ws.cell(i+1, 5, input_list[i][0])
        write_ws.cell(i+1, 6, input_list[i][1])

    write_wb.save(f'C:/Users/Hellllo/Desktop/크몽/쿠팡 크롤링/answer_excel_ddtester{excel_index}.xlsx')
    write_wb.close()
    excel_index += 1
    breaker +=1
