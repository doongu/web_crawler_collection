from openpyxl import load_workbook, Workbook
import crawler

#값 넣을 인자들 선언
check_name_list = [] #
excel_list = []
temp_excel_list = []
# temp_excel_list = []

#190*20 => 190, 20로 해서 나중에 190이 들어가는지, 20이 들어가는지 판별하기 위함.
def check_name_alg(input_row_c):

    global check_name_list

    for i in range(len(input_row_c)):
        # c에 * 이 있으면 *을 기준으로 나눠서 저장
        if input_row_c[i] == "*":
            check_name_list.append([input_row_c[:i], input_row_c[i + 1:]])
            break
    else:
        check_name_list.append([input_row_c])

#
def ignore_blank_alg(select_name, select_count, select_price):
    # 공백 무시 알고리즘
    # 이름이 빈칸이거나, 갯수가 빈칸이거나, 가격이 빈칸이면 실행
    global excel_list, temp_excel_list

    if ( select_name == "None" and select_count == "None" ) or select_price == "None":

        # temp_excel_list가 비어있지 않으면 그동안 추가한 데이터를 excel_list에 넣고 초기화 한다.
        if temp_excel_list:
            excel_list.append(temp_excel_list)
            temp_excel_list = []

    else:
        temp_excel_list.append([select_name, select_name + " " + select_count, int(select_price.replace(",", ""))])



def read_excel(excel_path):

    global excel_list, check_name_list

    # data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
    load_wb = load_workbook(f"{excel_path}", data_only=True)
    load_ws = load_wb['Sheet1']

    # 셀 주소로 값 출력
    for i in range(2 , 26):
        select_name = str(load_ws.cell(i, 2).value)#상품명
        select_count = str(load_ws.cell(i, 3).value) #수량
        select_price = str(load_ws.cell(i, 4).value) #가격

        #수량을 넣어서 check_name을 정한다.
        check_name_alg(select_count)

        #엑셀에서 공백이 있는데 이를 무시하고 값을 추가한다.
        ignore_blank_alg(select_name, select_count, select_price)

    return excel_list, check_name_list

#엑셀에 적는부분, crawler.select_data후 엑셀에 요소별로 저장한다.
def write_excel(excel_list, check_name_list):

    #모든 요소를 돌면 멈춘다. 첫 번째 "카테고리" 먼저 실행한다.
    for excel_index in range( len(excel_list) ):

        #가격을 리스트형태로 받는다. excel_list[excel_index] => excel_index 째 카테고리
        input_list = (crawler.select_data(excel_list[excel_index], check_name_list))

        #엑셀 열고
        write_wb =  Workbook()
        write_ws = write_wb.active

        for i in range(0, len(input_list)):
            write_ws.cell(i+1, 5, input_list[i][0]) # input_list[i][0]  i번째 제품 가격
            write_ws.cell(i+1, 6, input_list[i][1]) # input_list[i][1]  i번째 제품 url

        write_wb.save(f'C:/Users/Hellllo/Desktop/크몽/쿠팡 크롤링/coffee_6_14_ver_4{excel_index}.xlsx')
        write_wb.close()
        print(f"{excel_index}번째 엑셀을 저장했습니다.")

